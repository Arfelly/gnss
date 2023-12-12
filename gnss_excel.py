import os
from openpyxl import load_workbook

# Nombre del archivo Excel existente
archivo_excel_existente = 'V:/1DOCUMENTOS_MONITORES/CODIGOS/gnss/GNSS_Ovspop_POVASILP_prueba.xlsx'

# Carpeta que contiene los archivos .txt
carpeta_txt = 'V:/1DOCUMENTOS_MONITORES/CODIGOS/gnss/marco1/'

# Cargar el libro de trabajo existente
print(f'Ruta del archivo: {archivo_excel_existente}')
libro_existente = load_workbook(archivo_excel_existente)

print(libro_existente.sheetnames)

# Iterar sobre todos los archivos .txt en la carpeta
for nombre_txt in os.listdir(carpeta_txt):
    # Verificar si el archivo tiene la extensión .txt
    if nombre_txt.endswith('.txt'):
        # Obtener el nombre de la hoja correspondiente al archivo .txt
        nombre_hoja_destino = nombre_txt.split('.')[0]

        try:
            # Seleccionar la hoja de destino
            hoja_destino = libro_existente[nombre_hoja_destino]

            # Abrir el archivo .txt y leer los datos
            with open(os.path.join(carpeta_txt, nombre_txt), 'r') as archivo:
                lineas = archivo.readlines()

            # Procesar cada línea del archivo .txt
            for i, linea in enumerate(lineas, start=1):
                # Dividir la línea en columnas
                columnas = linea.strip().split('\t')  # Puedes ajustar el delimitador según el formato real

                # Escribir cada columna en la hoja de destino
                for j, columna in enumerate(columnas, start=1):
                    hoja_destino.cell(row=i, column=j, value=columna)

        except FileNotFoundError:
            print(f'Archivo {nombre_txt} no encontrado. Continuando con el siguiente archivo.')

# Guardar los cambios en el archivo Excel existente
libro_existente.save(archivo_excel_existente)

print(f'Datos agregados a las hojas correspondientes en el archivo {archivo_excel_existente}')

